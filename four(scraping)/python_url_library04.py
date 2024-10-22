#import time
import requests #->외부 라이브러리
#import urllib.request as req
from bs4 import BeautifulSoup
from lxml import etree
import openpyxl as op

wb = op.Workbook()
print(wb.sheetnames, "현 엑셀파일의 시트명 리스트")
ws = wb.active
ws.title="한빛 도서 목록"
#print(ws.title)
titles ="일련번호 도서명 저자 링크 가격".split()
ws.append(titles)


response = requests.get("https://www.hanbit.co.kr/media/books/category_list.html?cate_cd=001001")
#response = req.urlopen("https://www.hanbit.co.kr/media/books/category_list.html?cate_cd=001001").read()
soup = BeautifulSoup(response.content, 'html.parser')
#soup = BeautifulSoup(response, 'html.parser')
#book_list_first = soup.select_one("div.sub_book_list_area>ul>li.sub_book_list")
#body = soup.find("body")
#time.sleep(5) 5초 후 실행
#with open("abcdegh.html","w",encoding="utf-8") as f:
#    f.write(str(body))
#print("&&==================================잠자고 난 후 출력=====================================&&")
#book_writer = book_list_first.select_one("p.book_writer").text
#book_title = book_list_first.select_one("p.book_tit").text
#book_price = book_list_first.select_one("span.price").text
#print("제목 : ",book_title, "\n저자 : ",book_writer, "\n가격 : ",book_price)

print("책 리스트 목록 다 가져오기")
print("================================================================")

book_list_all = soup.select("div.sub_book_list_area>ul>li.sub_book_list")
for i in range(len(book_list_all)):
    append_data = []
    append_data.append(i+1)
    book_list = book_list_all[i]
    append_data.append( book_list.select_one("p.book_tit").text )
    append_data.append( book_list.select_one("p.book_writer").text )
    append_data.append( "https://www.hanbit.co.kr/media/books/" + book_list.select_one("p.book_tit>a")["href"][1:] )
    append_data.append( book_list.select_one("span.price").text[1:])#.replace(",","") 
#    print(book_list,book_title,book_writer,book_link,book_price,sep=" | ") # 이 부분 대신에 엑셀에 저장하는 작업
    ws.append(append_data)
    print(append_data)

for i in range(20):
    position = "E"+ str(i+1)
    ws[position].number_format = '₩#,##0'
wb.save("book_list.xlsx")
#print(book_list_all[0],len(book_list_all),sep="==================\n")
#dom = etree.HTML(str(body)) 
#xpath_str = '//*[@id="current-weather"]/div[4]/p' 
#xpath_str = '//*[@id="content"]/section/div/div[1]/p' 
#result = dom.xpath(xpath_str)
#print(result)
#print(dom.xpath(xpath_str)[0].text)
#웹페이지 실행시 자바 스크립트로 동적으로 변하기 때문에 null값을 반환한다. 결과가 제대로 나오지 않음