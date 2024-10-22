import openpyxl.worksheet
import pyautogui 
from bs4 import BeautifulSoup
import openpyxl 
import requests
from datetime import datetime

today=datetime.today().strftime("%Y%m%d")

pyautogui.hotkey("win","r")
pyautogui.sleep(2)
writeUrl="chrome https://finance.naver.com/sise/sise_quant.naver"
pyautogui.write(writeUrl)
pyautogui.sleep(2)
pyautogui.press("enter")


openEx = openpyxl.open(filename="D:/stockData/stock"+today+".xlsx")
activeSheet = openEx.active

url = "https://finance.naver.com/sise/sise_quant.naver"
response = requests.get(url)
soup = BeautifulSoup(response.content,'html.parser')



stock_list_all = soup.select("#contentarea > div.box_type_l > table > tr")
level = 10 # 10순위
for i in range(2,len(stock_list_all)): 
    append_data_arr=[]
    stock_list = stock_list_all[i].select("td")
    for k in range(len(stock_list)):
        append_data_arr.append(stock_list[k].text)
        if len(append_data_arr)==12 :
            activeSheet.append(append_data_arr)
        elif activeSheet.max_row == (level+2):
            break

max_num=activeSheet.max_row
max_colnum=activeSheet.max_column

column_arr = ['A','B','C','D','E','F','G','H','I','J','K']

sum_arr = []

for j in range(0,len(column_arr)):
    col_alphbet = column_arr[j]
    sum_func_arr = f"SUM("+column_arr[j]+"2:"+column_arr[j]+str(max_num)+")"
    #sum_arr.append(sum_func_arr)
    cellName=column_arr[j]+str(max_num)
    activeSheet[str(cellName)] = f"= {sum_func_arr}"
    #activeSheet[str(cellName)]=f"SUM("+column_arr[j]+"2:"+column_arr[j]+str(max_num)+")"
    #result_cell = activeSheet[str(cellName)]
    #activeSheet.append(sum_arr)

openEx.save(filename="D:/stockData/stock"+today+".xlsx")

# 수식 입력
    
    